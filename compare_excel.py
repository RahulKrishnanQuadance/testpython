import pandas as pd
import re
import os
from openpyxl import load_workbook
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation

# === CONFIG ===
DECIMAL_PLACES = 5  # Change this to set rounding precision

def excel_range_to_indexes(range_str):
    """Convert Excel range like A1:D20 to 0-based row/col indexes."""
    def col_to_index(col_str):
        col_str = col_str.upper()
        result = 0
        for char in col_str:
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result - 1

    match = re.match(r"([A-Za-z]+)(\d+):([A-Za-z]+)(\d+)", range_str)
    if not match:
        raise ValueError("Invalid range format. Use like A1:D20")

    start_col, start_row, end_col, end_row = match.groups()
    start_row, end_row = int(start_row) - 1, int(end_row) - 1
    start_col, end_col = col_to_index(start_col), col_to_index(end_col)

    return start_row, end_row, start_col, end_col

def normalize_decimal(val):
    """Round float-like values to given decimal places using Decimal."""
    if val is None:
        return None
    try:
        if isinstance(val, float):
            # Avoid float precision tails
            val = format(val, '.15g')
        dec_val = Decimal(val)
        quantize_pattern = '0.' + '0' * DECIMAL_PLACES
        return dec_val.quantize(Decimal(quantize_pattern), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError):
        return str(val).strip().lower()

def get_excel_col_letter(col_idx):
    """Convert a 0-based column index to Excel column letters."""
    col_letter = ""
    col_idx += 1
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        col_letter = chr(65 + remainder) + col_letter
    return col_letter

# --- Main Program ---
file1 = input("Enter path for first Excel file: ").strip()
sheet1 = input("Enter sheet name for first file: ").strip()

file2 = input("Enter path for second Excel file: ").strip()
sheet2 = input("Enter sheet name for second file: ").strip()

cell_range = input("Enter cell range to compare (e.g. A1:D20): ").strip()

# File names for output column headers
file1_name = os.path.basename(file1)
file2_name = os.path.basename(file2)

# Convert range to indexes
start_row, end_row, start_col, end_col = excel_range_to_indexes(cell_range)

# Load Excel files with formulas evaluated (data_only=True)
wb1 = load_workbook(file1, data_only=True)
ws1 = wb1[sheet1]

wb2 = load_workbook(file2, data_only=True)
ws2 = wb2[sheet2]

# Collect mismatches
mismatches = []

for row in range(start_row, end_row + 1):
    for col in range(start_col, end_col + 1):
        excel_row = row + 1
        excel_col_letter = get_excel_col_letter(col)
        cell_ref = f"{excel_col_letter}{excel_row}"

        val1 = ws1[cell_ref].value
        val2 = ws2[cell_ref].value

        norm1 = normalize_decimal(val1)
        norm2 = normalize_decimal(val2)

        if norm1 != norm2:
            raw1 = str(val1) if val1 is not None else ""
            raw2 = str(val2) if val2 is not None else ""
            mismatches.append([cell_ref, raw1, raw2])

# Save mismatches to Excel
if mismatches:
    output_df = pd.DataFrame(mismatches, columns=["Cell", file1_name, file2_name])
    output_df.to_excel("compareout.xlsx", index=False)
    print("✅ Mismatches found and saved to compareout.xlsx")
else:
    # Still create a file with a "no mismatches" message
    output_df = pd.DataFrame([["-", "No mismatches found", "No mismatches found"]],
                             columns=["Cell", file1_name, file2_name])
    output_df.to_excel("compareout.xlsx", index=False)
    print("✅ No mismatches found! compareout.xlsx created.")
